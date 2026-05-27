"""
modules/reporter.py — Styled Excel report generation with openpyxl.

Produces output/results.xlsx with:
  - "Results" sheet: styled data table with colour-coded success column
  - "Summary" sheet: totals, success rate, and execution timestamp
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from modules.logger import get_logger

logger = get_logger(__name__)

# ── Colour constants ──────────────────────────────────────────────────────────
_HEADER_BG   = "2E4057"   # dark blue-grey
_ROW_ALT     = "F0F4F8"   # light blue-grey for alternating rows
_SUCCESS_BG  = "C6EFCE"   # green tint
_FAILURE_BG  = "FFC7CE"   # red tint

# Column order must match the dicts produced by cmd_workflow
_FIELDNAMES = ["timestamp", "url", "status", "success", "error_message"]
_HEADERS    = ["Timestamp", "URL", "Status", "Success", "Error Message"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


# ── Public API ────────────────────────────────────────────────────────────────

def save_excel_report(results: list[dict[str, Any]]) -> Path:
    """Save *results* as a professionally styled Excel workbook.

    Args:
        results: List of result dicts produced during a workflow run.
                 Expected keys: timestamp, url, status, success, error_message.

    Returns:
        :class:`Path` to the saved ``results.xlsx`` file.
    """
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = config.OUTPUT_DIR / "results.xlsx"

    wb = openpyxl.Workbook()

    # ── Results sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Results"

    header_fill  = PatternFill("solid", fgColor=_HEADER_BG)
    header_font  = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    data_align   = Alignment(horizontal="left",   vertical="center")
    alt_fill     = PatternFill("solid", fgColor=_ROW_ALT)
    white_fill   = PatternFill("solid", fgColor="FFFFFF")
    success_fill = PatternFill("solid", fgColor=_SUCCESS_BG)
    failure_fill = PatternFill("solid", fgColor=_FAILURE_BG)
    border       = _thin_border()

    # Header row
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border

    # Data rows
    for row_idx, record in enumerate(results, start=2):
        is_alt = row_idx % 2 == 0
        base_fill = alt_fill if is_alt else white_fill

        for col_idx, field in enumerate(_FIELDNAMES, start=1):
            raw_value = record.get(field, "")

            if field == "success":
                is_success = bool(raw_value)
                value = "✓ Success" if is_success else "✗ Failed"
                fill  = success_fill if is_success else failure_fill
            else:
                value = raw_value
                fill  = base_fill

            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.alignment = data_align
            cell.border    = border

    # Auto-fit column widths (capped at 60 chars)
    for col_idx, header in enumerate(_HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, len(results) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")

    total     = len(results)
    successes = sum(1 for r in results if r.get("success"))
    failures  = total - successes
    rate      = (successes / total * 100) if total else 0.0
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    summary_rows = [
        ("Total Rows Processed", total),
        ("Total Successes",      successes),
        ("Total Failures",       failures),
        ("Success Rate (%)",     f"{rate:.1f}%"),
        ("Execution Timestamp",  timestamp),
    ]

    label_font = Font(bold=True)
    data_align = Alignment(horizontal="left", vertical="center")
    border     = _thin_border()

    for row_idx, (label, value) in enumerate(summary_rows, start=1):
        lc = ws_sum.cell(row=row_idx, column=1, value=label)
        lc.font      = label_font
        lc.alignment = data_align
        lc.border    = border

        vc = ws_sum.cell(row=row_idx, column=2, value=value)
        vc.alignment = data_align
        vc.border    = border

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 20

    wb.save(output_path)
    logger.info(f"Excel report saved → {output_path}")
    return output_path
